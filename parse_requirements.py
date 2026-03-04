import openpyxl
import json

def parse_excel(file_path):
    """解析Excel文件并输出内容"""
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        result = {
            "工作表列表": workbook.sheetnames,
            "数据": {}
        }
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n{'='*50}")
            print(f"工作表: {sheet_name}")
            print(f"{'='*50}")
            
            sheet_data = []
            
            # 读取所有行
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # 过滤空行
                if any(cell is not None for cell in row):
                    print(f"第{row_idx}行: {row}")
                    sheet_data.append(row)
            
            result["数据"][sheet_name] = sheet_data
        
        # 保存为JSON文件
        with open('需求解析结果.json', 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n{'='*50}")
        print("解析完成！结果已保存到 '需求解析结果.json'")
        print(f"{'='*50}")
        
        return result
        
    except Exception as e:
        print(f"解析出错: {str(e)}")
        return None

if __name__ == "__main__":
    parse_excel("需求.xlsx")
